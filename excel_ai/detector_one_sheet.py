import openpyxl
from .model import get_model
from .feature_engineering import extract_row_features_from_row
import pandas as pd 
from .feature_engineering import FEATURE_NAMES
from .config import DEFAULT_THRESHOLD, MAX_SCAN_ROWS

def detect_single_sheet(filepath: str, sheet_index: int = 0, threshold: float = DEFAULT_THRESHOLD):
    """
    Detect header row for a single sheet in an Excel file.

    Args:
        filepath (str): Path to Excel file
        sheet_index (int): Sheet index (0-based)
        threshold (float): Minimum probability threshold

    Returns:
        int | None: Detected header row number
    """

    model = get_model()
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    sheet_name = wb.sheetnames[sheet_index]
    ws = wb[sheet_name]

    total_cols = ws.max_column
    candidates = []

    rows = list(ws.iter_rows(max_row=MAX_SCAN_ROWS))

    for row_idx, row in enumerate(rows, start=1):
        features = extract_row_features_from_row(
            row, total_cols, ws, row_idx
        )

        if features is not None:

            features_df = pd.DataFrame([features], columns=FEATURE_NAMES)
            proba = model.predict_proba(features_df)[0][1]

            candidates.append((row_idx, proba))

    if not candidates:
        return None

    best_row, best_proba = max(candidates, key=lambda x: x[1])

    print(sheet_name, best_row, best_proba)

    return best_row
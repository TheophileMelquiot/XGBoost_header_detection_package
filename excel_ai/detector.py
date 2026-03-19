# excel_ai/detector.py

import openpyxl
from .model import get_model
from .feature_engineering import extract_row_features_from_row
import pandas as pd 
from .feature_engineering import FEATURE_NAMES
from .config import DEFAULT_THRESHOLD, MAX_SCAN_ROWS

model = get_model()

def detect_headers(filepath: str, threshold: float = DEFAULT_THRESHOLD):
    """
    Detect header rows for all sheets in an Excel file.

    Args:
        filepath (str): Path to Excel file
        threshold (float): Minimum probability to accept header

    Returns:
        dict: {sheet_name: header_row}
    """


    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    results = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total_cols = ws.max_column

        candidates = []

        rows = list(ws.iter_rows(max_row=MAX_SCAN_ROWS))

        for row_idx, row in enumerate(rows, start=1):
            features = extract_row_features_from_row(
                row, total_cols, ws, row_idx
            )

            if features is not None:
                
                features_df = pd.DataFrame([features], columns=FEATURE_NAMES)
                proba = model.predict_proba(features_df)[0][1]
                candidates.append((row_idx, proba))

        if not candidates:
            continue

        # Take highest probability row
        best_row, best_proba = max(candidates, key=lambda x: x[1])

        print(sheet_name, best_row, best_proba)

        results[sheet_name] = {
            "row": best_row,
            "confidence": float(best_proba)
        }


    return results
import openpyxl
import pandas as pd
from .model import get_model
from .feature_engineering import extract_row_features_from_row
from .feature_engineering import FEATURE_NAMES
from .config import MAX_SCAN_ROWS


model = get_model()


# --------------------------------------------------
# Propagation des cellules fusionnées (fallback)
# --------------------------------------------------

def expand_merged_cells(values):

    expanded = []
    last_value = None

    for v in values:

        if v not in (None, "", "nan"):
            last_value = str(v).strip()

        expanded.append(last_value)

    return expanded


# --------------------------------------------------
# Expansion des cellules fusionnées avec info openpyxl
# --------------------------------------------------

def get_merged_ranges_for_row(ws, row_idx):
    """
    Returns list of (min_col, max_col, value) for merged ranges that
    include the given row.
    """
    result = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row_idx <= merged_range.max_row:
            cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            result.append((merged_range.min_col, merged_range.max_col, cell.value))
    return result


def expand_merged_cells_with_ws(row_values, ws, row_idx):
    """
    Propagates merged cell values across all columns they span, using
    openpyxl's merged_cells information.  Returns a list aligned with
    row_values (same length).
    """
    merged_fill = {}
    for min_col, max_col, value in get_merged_ranges_for_row(ws, row_idx):
        if value not in (None, "", "nan"):
            for col in range(min_col, max_col + 1):
                merged_fill[col] = str(value).strip()

    result = []
    for col_idx, v in enumerate(row_values, start=1):
        if col_idx in merged_fill:
            result.append(merged_fill[col_idx])
        elif v not in (None, "", "nan"):
            result.append(str(v).strip())
        else:
            result.append(None)

    return result


def row_has_merged_cells(ws, row_idx):
    """
    Returns True if the given row contains at least one merged range that
    spans more than one column.
    """
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row_idx <= merged_range.max_row
                and merged_range.min_col != merged_range.max_col):
            return True
    return False


# --------------------------------------------------
# Détection robuste de sous-header
# --------------------------------------------------

def detect_repeating_pattern(values):

    cleaned = []

    for v in values:

        if v in (None, "", "nan"):
            cleaned.append("EMPTY")
        else:
            cleaned.append(str(v).strip().lower())

    n = len(cleaned)

    if n < 4:
        return False

    # ratio valeurs uniques
    unique_ratio = len(set(cleaned)) / n

    counts = {}

    for v in cleaned:
        counts[v] = counts.get(v, 0) + 1

    max_repeat = max(counts.values())

    repetition_ratio = max_repeat / n

    if unique_ratio < 0.7 or repetition_ratio > 0.25:
        return True

    return False


# --------------------------------------------------
# Fusion de deux lignes header
# --------------------------------------------------

def is_generic_parent(parent_values):
    """
    Returns True if all non-empty parent values are the same string,
    indicating a single merged cell spanning multiple columns with no
    meaningful sub-column names (e.g. "tous clients" repeated 4 times).
    """
    non_empty = [v for v in parent_values if v not in (None, "", "nan")]
    if len(non_empty) <= 1:
        return False
    return len(set(non_empty)) == 1


def merge_headers(parent, child):
    """
    Merges parent and child header rows into a single list of column names.

    For merged parent cells (same value repeated across consecutive columns),
    generates numbered variants like "tous clients 1", "tous clients 2", …
    instead of concatenating the parent with actual data values.

    For single-column parents, uses the standard "parent child" concatenation.
    """
    merged = []
    n = len(parent)
    i = 0

    while i < n:
        p_val = parent[i]
        c_val = child[i] if i < len(child) else None

        if p_val in (None, "", "nan"):
            # No parent value – keep the child value as-is
            if c_val not in (None, "", "nan"):
                merged.append(str(c_val).strip())
            else:
                merged.append(None)
            i += 1
            continue

        # Determine how many consecutive columns share this parent value
        j = i + 1
        while j < n and parent[j] == p_val:
            j += 1

        if is_generic_parent(parent[i:j]):
            # Generic merged cell: produce numbered variants, ignore child
            label = str(p_val).strip()
            for k in range(j - i):
                merged.append(f"{label} {k + 1}")
        else:
            # Single column: standard parent + child concatenation
            parts = [str(p_val).strip()]
            if c_val not in (None, "", "nan"):
                parts.append(str(c_val).strip())
            merged.append(" ".join(parts))

        i = j

    return merged


# --------------------------------------------------
# Détection headers améliorée
# --------------------------------------------------

def detect_headers_upgrade(filepath):

    # read_only must be False to access ws.merged_cells
    wb = openpyxl.load_workbook(filepath, data_only=True)

    results = {}

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        total_cols = ws.max_column

        candidates = []

        rows = list(ws.iter_rows(max_row=MAX_SCAN_ROWS))

        # --------------------------------------------
        # Étape 1 : prédiction ML
        # --------------------------------------------

        for row_idx, row in enumerate(rows, start=1):

            features = extract_row_features_from_row(
                row,
                total_cols,
                ws,
                row_idx
            )

            if features is None:
                continue

            df = pd.DataFrame([features], columns=FEATURE_NAMES)

            proba = model.predict_proba(df)[0][1]

            candidates.append((row_idx, proba))

        if not candidates:
            continue

        best_row, best_proba = max(candidates, key=lambda x: x[1])

        # --------------------------------------------
        # Étape 2 : analyse du header détecté
        # --------------------------------------------

        raw_header_values = [c.value for c in ws[best_row]]

        # Expand merged cells using openpyxl info before filtering
        header_values = expand_merged_cells_with_ws(raw_header_values, ws, best_row)

        # Non-empty values for pattern analysis
        non_empty_header = [v for v in header_values if v is not None]

        reconstructed_header = None
        header_rows = [best_row]

        # --------------------------------------------
        # Étape 3 : détecter sous-header
        # --------------------------------------------

        if best_row > 1:
            parent_row_idx = best_row - 1

            # Trigger parent merge if the parent row has multi-column merged
            # cells (group labels like "tous clients") OR if the current
            # header row itself shows a repeating sub-header pattern.
            parent_has_merged = row_has_merged_cells(ws, parent_row_idx)
            header_is_repeating = detect_repeating_pattern(non_empty_header)

            if parent_has_merged or header_is_repeating:

                raw_parent = [c.value for c in ws[parent_row_idx]]

                parent_values = expand_merged_cells_with_ws(
                    raw_parent, ws, parent_row_idx
                )

                # Fallback: simple forward-fill if ws info yields nothing
                if all(v is None for v in parent_values):
                    parent_values = expand_merged_cells(raw_parent)

                reconstructed_header = merge_headers(parent_values, header_values)

                # Drop fully-empty merged columns
                reconstructed_header = [
                    v for v in reconstructed_header if v is not None
                ]

                header_rows = [parent_row_idx, best_row]

        # --------------------------------------------
        # Résultat
        # --------------------------------------------

        results[sheet_name] = {
            "header_rows": header_rows,
            "confidence": float(best_proba),
            "columns": reconstructed_header if reconstructed_header else non_empty_header
        }

    return results